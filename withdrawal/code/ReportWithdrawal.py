import openpyxl
import yaml


class ReportWithdrawal:
    config_path = "/path_to/config.yaml"
    config_withdrawal = "/path_to/withdrawal.yaml"

    def __init__(self):
        # get settings from config.yaml
        with open(self.config_path, "r") as file:
            data = yaml.safe_load(file)
            columns = data["ozon"]["settings"]["columns"]
            self.id_column = columns["id"]
            self.vendor_code_column = columns["vendor_code"]
            self.size_column = columns["size"]
            self.cis_column = columns["cis"]
            self.sending_id_column = columns["sending_id"]
            self.sale_price_column = columns["sale_price"]
            self.payroll_column = columns["payroll"]
            self.act_of_sale_id_column = columns["act_of_sale_id"]
            self.data_sale_column = columns["data_sale"]
            self.act_of_return_id_column = columns["act_of_return_id"]
            self.data_return_column = columns["data_return"]

        # get settings from withdrawal.yaml
        with open(self.config_withdrawal, "r") as file:
            data = yaml.safe_load(file)
            withdrawal_settings = data["ozon"]["withdrawal"]
            self.inn = str(withdrawal_settings["inn"])
            self.document_path = withdrawal_settings["document_path"]
            self.month = withdrawal_settings["month"]
            self.data = withdrawal_settings["data"] + " 00:00:00"

        # open book and sheet xlsx
        self.workbook = openpyxl.load_workbook(self.document_path)
        self.sheet = self.workbook[self.month]

    def get_xml_cis(self, cis):
        cis = cis.replace("&", "&amp;")
        cis = cis.replace("<", "&lt;")
        cis = cis.replace(">", "&gt;")
        cis = cis.replace("\"", "&quot;")
        cis = cis.replace("\'", "&apos;")
        return cis

    def get_xml_product(self, cis, price):
        cis = self.get_xml_cis(cis)
        product = "\t\t<product>\n"
        product += "\t\t\t<cis>" + cis + "</cis>\n"
        product += "\t\t\t<cost>" + price + "00" + "</cost>\n"
        product += "\t\t</product>\n"
        return product

    def get_xml_header(self, inn, data, act_of_sale_number):
        header = "<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n" + \
                 "<withdrawal version=\"8\">\n" + \
                 "\t<trade_participant_inn>" + inn + "</trade_participant_inn>\n" + \
                 "\t<withdrawal_type>DISTANCE</withdrawal_type>\n" + \
                 "\t<withdrawal_date>" + data + "</withdrawal_date>\n" + \
                 "\t<primary_document_type>OTHER</primary_document_type>\n" + \
                 "\t<primary_document_number>" + act_of_sale_number + "</primary_document_number>\n" + \
                 "\t<primary_document_date>" + data + "</primary_document_date>\n" + \
                 "\t<primary_document_custom_name>АКТПРИЕМАПЕРЕДАЧИ</primary_document_custom_name>\n"
        return header

    def get_xml_report(self):
        # формируем список id у которых в строке есть значение в поле дата равно значению даты пользователя
        list_of_ids = []
        for i in range(2, self.sheet.max_row + 1):
            tmp = self.sheet.cell(row=i, column=self.data_sale_column)
            if str(tmp.value) == self.data:
                list_of_ids.append(i)

        # получаем номер акта прием-передачи
        act_of_sale_number = str(self.sheet.cell(row=list_of_ids[0], column=self.act_of_sale_id_column).value)
        act_of_sale_number = " ".join(act_of_sale_number.split())

        # формируем заголовок файла
        answer = self.get_xml_header(self.inn, self.data, act_of_sale_number)

        # формируем список xml марок и их цен для каждого товара
        answer += "\t<products_list>\n"
        for i in list_of_ids:
            cis = str(self.sheet.cell(row=i, column=self.cis_column).value)
            price = str(self.sheet.cell(row=i, column=self.sale_price_column).value)
            answer += self.get_xml_product(cis, price)
        answer += "\t</products_list>\n</withdrawal>"

        # формируем файл и сохраняем
        file_out = open("reports/"+self.data + ".xml", "a")
        file_out.write(answer)
        file_out.close()
